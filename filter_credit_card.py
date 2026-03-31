import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def append_df_to_xlsx(path, df, write_header=False):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        if write_header:
            ws.append(list(df.columns))
        for r in df.itertuples(index=False, name=None):
            ws.append(list(r))
        wb.save(path)
    else:
        wb = load_workbook(path)
        ws = wb.active
        if write_header:
            ws.append(list(df.columns))
        for r in df.itertuples(index=False, name=None):
            ws.append(list(r))
        wb.save(path)

def filter_chunk(df, colname='Information Type Name'):
    # attempt to normalize column name spacing
    if colname not in df.columns:
        cols_map = {c.strip(): c for c in df.columns}
        if colname in cols_map:
            df = df.rename(columns={cols_map[colname]: colname})
        else:
            raise KeyError(f"Column '{colname}' not found")
    s = df[colname].fillna('').astype(str)
    mask = s.str.contains('credit card', case=False, na=False)
    return df[mask], len(df), mask.sum()

def main():
    clear_screen()
    fn = input("Enter CSV file name (in same folder): ").strip()
    if not os.path.exists(fn):
        print("File not found:", fn); return

    clear_screen()
    start = time.perf_counter()
    base, _ = os.path.splitext(os.path.basename(fn))
    out_fn = os.path.join(os.path.dirname(os.path.abspath(fn)), f"{base}_output.xlsx")

    chunksize = 200_000
    first_write = True
    total_rows = total_matched = 0

    try:
        reader = pd.read_csv(fn, chunksize=chunksize, dtype=str, low_memory=False)
    except Exception as e:
        print("Error opening CSV:", e); return

    try:
        for df in reader:
            matched_df, rows_in_chunk, matched_in_chunk = filter_chunk(df)
            total_rows += rows_in_chunk
            if matched_in_chunk:
                append_df_to_xlsx(out_fn, matched_df, write_header=first_write)
                first_write = False
                total_matched += matched_in_chunk
    except KeyError as ke:
        print("Error:", ke); return
    except Exception as e:
        print("Processing error:", e); return

    elapsed = time.perf_counter() - start
    discarded = total_rows - total_matched

    print("Output file:", out_fn)
    print("Rows processed:", total_rows)
    print("Records with 'Credit Card':", total_matched)
    print("Records discarded:", discarded)
    print(f"Total execution time: {elapsed:.2f} seconds")

if __name__ == '__main__':
    main()
